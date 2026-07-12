#!/usr/bin/env python3
"""import_testcases_excel.py — Excel test cases → canonical JSON
(the reverse of ../../qa-agent/scripts/export_json_to_excel.py).

Reads the first worksheet, matches columns BY HEADER NAME (the exporter's open
columns AND the review-table variants; extra columns are ignored, missing ones
reported), and writes the canonical test-case JSON the gen-auto-test / qa-agent
flows consume (references/json-contract.md).

Needs openpyxl (not a framework dependency):

    uv run --with openpyxl python3 \
      .claude/skills/gen-auto-test/scripts/import_testcases_excel.py \
      --xlsx TestCases_search.xlsx \
      --out  test-output/ai/testcases-search.json \
      [--feature "Search"] [--story EAST-123]

Step parsing: the Steps cell is split back into stepDetails — one step per
line, "N. action | element: <selector>" (the exporter's format) or plain
"N. action" / free lines.

Exit codes: 0 = written, 2 = usage / file error.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

#: Header text (lowercased, trimmed) → canonical field. Covers the exporter's
#: open columns (ID, Title, …) and the review-table variants (TC ID, Pr., …).
HEADER_MAP = {
    "id": "tcId",
    "tc id": "tcId",
    "title": "testDescription",
    "test description": "testDescription",
    "feature": "feature",
    "sub-feature": "subFeature",
    "preconditions": "summaryPrecondition",
    "summary & specific pre-condition": "summaryPrecondition",
    "summary & pre-condition": "summaryPrecondition",
    "summary": "summaryPrecondition",
    "steps": "stepDetails",
    "step details": "stepDetails",
    "expected result": "expectedResult",
    "priority": "priority",
    "pr.": "priority",
    "type": "type",
    "ac": "acIds",
    "status": "testResult",
    "test result": "testResult",
    "bug id": "bugId",
    "notes": "notes",
}


def parse_steps(cell_text: str) -> list[dict]:
    if not cell_text:
        return []
    steps = []
    for index, raw in enumerate(str(cell_text).splitlines()):
        line = raw.strip()
        if not line:
            continue
        number = None
        body = line
        head, dot, rest = line.partition(".")
        if dot and head.strip().isdigit():
            number, body = int(head), rest.strip()
        detail, sep, element = body.partition("| element:")
        steps.append(
            {
                "step": number if number is not None else index + 1,
                "detail": detail.strip() or body,
                "element": element.strip() if sep else "",
            }
        )
    return steps


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel test cases -> canonical JSON.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--feature", default="")
    parser.add_argument("--story", default="")
    args = parser.parse_args()

    xlsx = Path(args.xlsx).resolve()
    if not xlsx.is_file():
        print(f"Error: file not found: {xlsx}", file=sys.stderr)
        return 2

    ws = load_workbook(xlsx, read_only=True, data_only=True).worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        print("Error: worksheet is empty.", file=sys.stderr)
        return 2

    columns: dict[str, int] = {}
    for idx, cell in enumerate(header):
        key = HEADER_MAP.get(str(cell or "").strip().lower())
        if key and key not in columns:
            columns[key] = idx

    missing = [k for k in ("tcId", "stepDetails") if k not in columns]
    if missing:
        found = " | ".join(str(c or "") for c in header)
        print(
            f"Error: required column(s) not found by header: {', '.join(missing)}. "
            f"Found headers: {found}",
            file=sys.stderr,
        )
        return 2

    test_cases = []
    for row in rows:
        def get(key: str, row: tuple = row) -> str:
            idx = columns.get(key)
            return str(row[idx] or "").strip() if idx is not None and idx < len(row) else ""

        tc_id = get("tcId")
        if not tc_id:
            continue  # skip empty rows
        automatable = get("type").lower() != "manual"
        test_cases.append(
            {
                "tcId": tc_id,
                "feature": get("feature") or args.feature,
                "subFeature": get("subFeature"),
                "summaryPrecondition": get("summaryPrecondition"),
                "testDescription": get("testDescription"),
                "expectedResult": get("expectedResult"),
                "stepDetails": parse_steps(get("stepDetails")),
                "priority": get("priority") or "P2",
                "priorityReason": "",
                "duplicateStatus": "",
                "duplicateOf": "",
                "duplicateReason": "",
                "acIds": [a.strip() for a in get("acIds").split(",") if a.strip()],
                # The agent reviews/infers these in Phase 1 — imported sheets rarely carry them.
                "automatable": automatable,
                "surface": "ui",
                "coverageType": "",
                "specFile": "",
                "xrayKey": "",
                "testrailCaseId": "",
                "testResult": get("testResult"),
                "bugId": get("bugId"),
                "notes": get("notes"),
            }
        )

    if not test_cases:
        print("Error: no test-case rows found under the header row.", file=sys.stderr)
        return 2

    doc = {
        "schemaVersion": "aiqa.qa.excel.v1",
        "meta": {
            "feature": args.feature or test_cases[0]["feature"],
            "userStoryKey": args.story,
            "sourceNoteTitle": "",
            "figmaDesignLink": "",
            "prototypeLink": "",
            "generatedAt": datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "approvalRequired": True,
            "approvalStatus": "draft",
            "testManagement": "excel",
            "testExecutionKey": "",
            "testrailRunId": "",
            "importedFrom": str(xlsx),
        },
        "testCases": test_cases,
        "assumptions": [],
        "openQuestions": [],
    }

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"[import-excel] {len(test_cases)} case(s) -> {out}")
    print(
        "[import-excel] review needed: automatable/surface/coverageType default to "
        "inferred/ui/(empty) — the agent infers and marks them in Phase 1."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
