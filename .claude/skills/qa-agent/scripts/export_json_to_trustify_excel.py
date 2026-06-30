#!/usr/bin/env python3
"""export_json_to_trustify_excel.py — approved QA JSON -> Trustify-style Excel.

STEP 13 of the workflow. Reads the approved canonical JSON (see
references/json-contract.md) and writes an .xlsx in the Trustify test-case format.

Needs openpyxl, which is NOT a framework dependency — run it with an ephemeral
dependency so pyproject.toml (patch-guarded) is untouched:

    uv run --with openpyxl python3 \
      .claude/skills/qa-agent/scripts/export_json_to_trustify_excel.py \
      --json docs/ai/testcases/TestCases_<feature>.json \
      --out test-output/qa/TestCases_<feature>.xlsx

Pass --template <file.xlsx> to fill an existing template instead of a fresh book.
Export only AFTER the exact approval phrase `I approve`.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

EXPECTED_HEADERS = [
    "TC ID",
    "Feature",
    "Sub-feature",
    "Summary & Specific pre-condition",
    "Test Description",
    "Pr.",
    "Test Result",
    "Bug ID",
    "Notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export QA JSON to Trustify-style Excel.")
    parser.add_argument("--json", required=True, help="Path to approved JSON file.")
    parser.add_argument("--out", required=True, help="Path to output XLSX file.")
    parser.add_argument("--template", required=False, help="Optional XLSX template path.")
    return parser.parse_args()


def load_json(path: str) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as f:
        return json.load(f)


def ensure_workbook(template_path: str | None):
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"  # type: ignore
    ws.append(EXPECTED_HEADERS)  # type: ignore
    return wb, ws


def ensure_headers(ws) -> None:  # type: ignore
    current = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    if current != EXPECTED_HEADERS:
        for i, header in enumerate(EXPECTED_HEADERS, start=1):
            ws.cell(row=1, column=i, value=header)


def stringify_step_details(step_details: Any) -> str:
    if not step_details:
        return ""
    lines: list[str] = []
    if isinstance(step_details, list):
        for item in step_details:
            if isinstance(item, dict):
                step = item.get("step", "")
                detail = item.get("detail", "")
                element = item.get("element", "")
                text = f"{step}. {detail}".strip()
                if element:
                    text += f" | element: {element}"
                lines.append(text)
            else:
                lines.append(str(item))
    else:
        lines.append(str(step_details))
    return "\n".join(lines)


def build_test_description(tc: dict[str, Any]) -> str:
    value = tc.get("testDescription", "")
    step_text = stringify_step_details(tc.get("stepDetails"))
    if value and step_text:
        return f"{value}\n\nStep details:\n{step_text}"
    if step_text:
        return step_text
    return str(value) if value is not None else ""


def build_notes(tc: dict[str, Any]) -> str:
    note_parts = []
    if tc.get("notes"):
        note_parts.append(str(tc["notes"]))
    if tc.get("priorityReason"):
        note_parts.append(f"Priority reason: {tc['priorityReason']}")
    if tc.get("duplicateStatus"):
        note_parts.append(f"Duplicate status: {tc['duplicateStatus']}")
    if tc.get("duplicateOf"):
        note_parts.append(f"Duplicate of: {tc['duplicateOf']}")
    if tc.get("duplicateReason"):
        note_parts.append(f"Duplicate reason: {tc['duplicateReason']}")
    return "\n".join(note_parts)


def export(data: dict[str, Any], out_path: str, template_path: str | None) -> None:
    wb, ws = ensure_workbook(template_path)
    ensure_headers(ws)

    if ws is None:
        raise ValueError("Failed to create or load workbook/worksheet.")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for tc in data.get("testCases", []):
        ws.append(
            [
                tc.get("tcId", ""),
                tc.get("feature", data.get("meta", {}).get("feature", "")),
                tc.get("subFeature", ""),
                tc.get("summaryPrecondition", ""),
                build_test_description(tc),
                str(tc.get("priority", "")),
                tc.get("testResult", ""),
                tc.get("bugId", ""),
                build_notes(tc),
            ]
        )

    Path(out_path).resolve().parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    args = parse_args()
    data = load_json(args.json)
    export(data, args.out, args.template)
    print(f"✅ Exported {len(data.get('testCases', []))} test case(s) -> {args.out}")


if __name__ == "__main__":
    main()
