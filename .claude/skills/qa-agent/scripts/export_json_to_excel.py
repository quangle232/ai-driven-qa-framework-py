#!/usr/bin/env python3
"""export_json_to_excel.py — approved QA JSON -> open, tool-agnostic Excel.

STEP 13 (Excel target) of the workflow. Reads the approved canonical JSON (see
references/json-contract.md) and writes an .xlsx in an OPEN test-case format whose
columns import cleanly into Xray, TestRail, Zephyr, etc. — the neutral interchange
the client can take into whichever test-management tool they choose.

Needs openpyxl, which is NOT a framework dependency — run it with an ephemeral
dependency so pyproject.toml (patch-guarded) is untouched:

    uv run --with openpyxl python3 \
      .claude/skills/qa-agent/scripts/export_json_to_excel.py \
      --json docs/ai/testcases/TestCases_<feature>.json \
      --out test-output/qa/TestCases_<feature>.xlsx

Pass --template <file.xlsx> to fill a tool-specific template instead of a fresh
book. Export only AFTER the exact approval phrase `I approve`.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# Open, tool-agnostic header set (imports cleanly into Xray / TestRail / Zephyr).
EXPECTED_HEADERS = [
    "ID",
    "Title",
    "Feature",
    "Sub-feature",
    "Preconditions",
    "Steps",
    "Expected Result",
    "Priority",
    "Type",
    "AC",
    "Status",
    "Bug ID",
    "Notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export QA JSON to an open Excel format.")
    parser.add_argument("--json", required=True, help="Path to approved JSON file.")
    parser.add_argument("--out", required=True, help="Path to output XLSX file.")
    parser.add_argument("--template", required=False, help="Optional XLSX template path.")
    return parser.parse_args()


def load_json(path: str) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as f:
        return json.load(f)


def ensure_workbook(template_path: str | None):
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"  # type: ignore
    ws.append(EXPECTED_HEADERS)  # type: ignore
    return wb, ws


def ensure_headers(ws) -> None:  # type: ignore
    current = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    if current != EXPECTED_HEADERS:
        for i, header in enumerate(EXPECTED_HEADERS, start=1):
            ws.cell(row=1, column=i, value=header)


def stringify_step_details(step_details: Any) -> str:
    if not step_details:
        return ""
    lines: list[str] = []
    if isinstance(step_details, list):
        for item in step_details:
            if isinstance(item, dict):
                step = item.get("step", "")
                detail = item.get("detail", "")
                element = item.get("element", "")
                text = f"{step}. {detail}".strip()
                if element:
                    text += f" | element: {element}"
                lines.append(text)
            else:
                lines.append(str(item))
    else:
        lines.append(str(step_details))
    return "\n".join(lines)


def build_notes(tc: dict[str, Any]) -> str:
    note_parts = []
    if tc.get("notes"):
        note_parts.append(str(tc["notes"]))
    if tc.get("priorityReason"):
        note_parts.append(f"Priority reason: {tc['priorityReason']}")
    if tc.get("duplicateStatus"):
        note_parts.append(f"Duplicate status: {tc['duplicateStatus']}")
    if tc.get("duplicateOf"):
        note_parts.append(f"Duplicate of: {tc['duplicateOf']}")
    if tc.get("duplicateReason"):
        note_parts.append(f"Duplicate reason: {tc['duplicateReason']}")
    return "\n".join(note_parts)


def row_for(tc: dict[str, Any], default_feature: str) -> list[str]:
    tc_type = "Automation" if tc.get("automatable") else "Manual"
    return [
        tc.get("tcId", ""),
        tc.get("testDescription", ""),
        tc.get("feature", default_feature),
        tc.get("subFeature", ""),
        tc.get("summaryPrecondition", ""),
        stringify_step_details(tc.get("stepDetails")),
        tc.get("expectedResult", ""),
        str(tc.get("priority", "")),
        tc_type,
        ", ".join(tc.get("acIds", []) or []),
        tc.get("testResult", ""),
        tc.get("bugId", ""),
        build_notes(tc),
    ]


def export(data: dict[str, Any], out_path: str, template_path: str | None) -> int:
    wb, ws = ensure_workbook(template_path)
    ensure_headers(ws)

    if ws is None:
        raise ValueError("Failed to create or load workbook/worksheet.")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    default_feature = data.get("meta", {}).get("feature", "")
    cases = data.get("testCases", [])
    for tc in cases:
        ws.append(row_for(tc, default_feature))

    Path(out_path).resolve().parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(cases)


def main() -> None:
    args = parse_args()
    data = load_json(args.json)
    count = export(data, args.out, args.template)
    print(f"✅ Exported {count} test case(s) -> {args.out}")


if __name__ == "__main__":
    main()
